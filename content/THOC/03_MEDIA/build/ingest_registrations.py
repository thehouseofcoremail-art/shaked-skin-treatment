# -*- coding: utf-8 -*-
"""
Read a THOC registrations export and work out what August's record actually is.

    python3 ingest_registrations.py <file.xlsx|file.csv> [--sheet NAME]

Prints what it found in the file first — sheets, headers, row count, date span —
then the per-month table, then the record checks. The description always works
even when the column guesses miss, so a mismatched file costs one look rather
than a round trip.

A "record" needs history. If the export only covers August there is nothing to
beat, and the script says so instead of quietly calling the first month a peak.

Writes august_stats.json for carousel.py to read.
"""
import sys, os, json, re, csv
from collections import defaultdict, Counter
from datetime import datetime, date

AUG = (2026, 8)

# Header guesses, widest first. Hebrew and English, since exports vary.
COLS = {
    "date":   ["תאריך", "תאריך שיעור", "תאריך הרשמה", "יום", "date", "class date",
               "booking date", "created", "datetime", "start"],
    "member": ["שם", "שם מלא", "מתאמן", "מתאמנת", "לקוח", "לקוחה", "member", "name",
               "client", "customer", "email", "מייל", "אימייל", "טלפון", "phone"],
    "cls":    ["שיעור", "סוג שיעור", "פורמט", "class", "class type", "service",
               "activity", "מקצוע"],
    "status": ["סטטוס", "status", "attended", "נוכחות", "booking status"],
}
CANCELLED = {"cancel", "cancelled", "canceled", "no show", "no-show", "בוטל",
             "ביטול", "לא הגיע", "לא הגיעה"}


def norm(s):
    return re.sub(r"\s+", " ", str(s or "")).strip().lower()


def load(path, sheet=None):
    """-> (headers, rows, source_label). Rows are lists aligned to headers."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".csv", ".tsv", ".txt"):
        delim = "\t" if ext == ".tsv" else ","
        with open(path, encoding="utf-8-sig", newline="") as fh:
            r = list(csv.reader(fh, delimiter=delim))
        return r[0], r[1:], os.path.basename(path)

    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    print(f"  sheets: {wb.sheetnames}")
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    # Skip any banner rows above the real header.
    hi = 0
    for i, r in enumerate(rows[:10]):
        if sum(1 for c in r if c not in (None, "")) >= max(2, len(r) // 3):
            hi = i
            break
    return rows[hi], rows[hi + 1:], f"{os.path.basename(path)} :: {ws.title}"


def find(headers, kind):
    hs = [norm(h) for h in headers]
    for want in COLS[kind]:
        for i, h in enumerate(hs):
            if h == want:
                return i
    for want in COLS[kind]:
        for i, h in enumerate(hs):
            if want and want in h:
                return i
    return None


def as_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date):     return v
    s = str(v or "").strip()
    if not s: return None
    for f in ("%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y", "%d-%m-%Y", "%m/%d/%Y",
              "%Y/%m/%d", "%d/%m/%y", "%d.%m.%y"):
        try: return datetime.strptime(s[:10], f).date()
        except ValueError: pass
    return None


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    path = sys.argv[1]
    sheet = None
    if "--sheet" in sys.argv:
        sheet = sys.argv[sys.argv.index("--sheet") + 1]
    if not os.path.exists(path):
        sys.exit(f"not found: {path}")

    print(f"\nFILE  {path}")
    headers, rows, label = load(path, sheet)
    print(f"  headers ({len(headers)}): {[str(h) for h in headers]}")
    print(f"  data rows: {len(rows)}")

    ci = {k: find(headers, k) for k in COLS}
    print("\nCOLUMN MATCH")
    for k, i in ci.items():
        print(f"  {k:7} -> {headers[i] if i is not None else '(not found)'}")
    if ci["date"] is None:
        sys.exit("\nNo date column found. Re-run with --sheet, or tell me which column holds the date.")

    per = defaultdict(lambda: {"regs": 0, "members": set(), "classes": set(),
                               "formats": Counter(), "days": Counter()})
    first_seen, bad, undated = {}, 0, 0
    for r in rows:
        d = as_date(r[ci["date"]]) if ci["date"] < len(r) else None
        if not d:
            undated += 1
            continue
        if ci["status"] is not None and ci["status"] < len(r):
            if any(c in norm(r[ci["status"]]) for c in CANCELLED):
                bad += 1
                continue
        m = per[(d.year, d.month)]
        m["regs"] += 1
        who = norm(r[ci["member"]]) if ci["member"] is not None and ci["member"] < len(r) else None
        if who:
            m["members"].add(who)
            if who not in first_seen or d < first_seen[who]:
                first_seen[who] = d
        fmt = str(r[ci["cls"]]).strip() if ci["cls"] is not None and ci["cls"] < len(r) else None
        if fmt and fmt.lower() != "none":
            m["formats"][fmt] += 1
            m["classes"].add((d, fmt))
        m["days"][d] += 1

    if not per:
        sys.exit("No dated rows parsed.")
    print(f"\n  skipped: {undated} undated, {bad} cancelled/no-show")

    months = sorted(per)
    print(f"\nPER MONTH  ({months[0][1]}/{months[0][0]} → {months[-1][1]}/{months[-1][0]})")
    print(f"  {'month':9}{'regs':>8}{'members':>9}{'new':>7}{'classes':>9}")
    table = {}
    for y, mo in months:
        v = per[(y, mo)]
        new = sum(1 for w, d0 in first_seen.items()
                  if (d0.year, d0.month) == (y, mo) and w in v["members"])
        table[f"{y}-{mo:02d}"] = {"regs": v["regs"], "members": len(v["members"]),
                                  "new": new, "classes": len(v["classes"])}
        print(f"  {y}-{mo:02d}  {v['regs']:>8}{len(v['members']):>9}{new:>7}{len(v['classes']):>9}")

    if AUG not in per:
        sys.exit(f"\nNo rows for {AUG[1]}/{AUG[0]} in this file.")

    a = table[f"{AUG[0]}-{AUG[1]:02d}"]
    prior = {k: v for k, v in table.items() if k < f"{AUG[0]}-{AUG[1]:02d}"}
    print("\nRECORD CHECK")
    records = {}
    if not prior:
        print("  No month precedes August in this file — a record cannot be claimed from it.")
        print("  Send an export covering earlier months, or the slide says a plain count.")
    else:
        for k in ("regs", "members", "new", "classes"):
            best = max(v[k] for v in prior.values())
            records[k] = a[k] > best
            mark = "RECORD" if a[k] > best else f"below peak {best}"
            print(f"  {k:8} august {a[k]:>6}   previous best {best:>6}   {mark}")

    v = per[AUG]
    top_fmt = v["formats"].most_common(3)
    top_day = v["days"].most_common(1)
    print("\nAUGUST DETAIL")
    print(f"  busiest formats: {top_fmt or '(no class column)'}")
    if top_day:
        print(f"  busiest day:     {top_day[0][0]}  ({top_day[0][1]} registrations)")

    out = {
        "source": label, "generated": datetime.now().isoformat(timespec="seconds"),
        "august": a, "per_month": table, "records": records,
        "has_history": bool(prior),
        "top_formats": top_fmt,
        "busiest_day": [str(top_day[0][0]), top_day[0][1]] if top_day else None,
    }
    dest = os.path.join(os.path.dirname(os.path.abspath(__file__)), "august_stats.json")
    with open(dest, "w", encoding="utf-8") as fh:
        json.dump(out, fh, ensure_ascii=False, indent=2)
    print(f"\nwrote {dest}")


if __name__ == "__main__":
    main()
